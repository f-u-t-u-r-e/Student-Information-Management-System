#!/usr/bin/env python3
import http.server
import socketserver
import json
import io
import os
from urllib.parse import urlparse, parse_qs

ROOT = os.path.dirname(__file__)
DATA_FILE = os.path.join(ROOT, 'data', 'students.json')
EXPORT_DIR = os.path.join(ROOT, 'exports')
EXPORT_METADATA = os.path.join(EXPORT_DIR, 'exports.json')

def ensure_data_dir():
    d = os.path.dirname(DATA_FILE)
    if not os.path.exists(d):
        os.makedirs(d)
    if not os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump([], f, ensure_ascii=False, indent=2)
    # ensure exports dir exists
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)
    # ensure exports metadata exists
    if not os.path.exists(EXPORT_METADATA):
        with open(EXPORT_METADATA, 'w', encoding='utf-8') as f:
            json.dump([], f, ensure_ascii=False, indent=2)

def load_students():
    ensure_data_dir()
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_students(lst):
    ensure_data_dir()
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(lst, f, ensure_ascii=False, indent=2)

def calc_gpa(courses):
    total1 = 0.0
    total2 = 0.0
    for c in courses:
        total1 += c.get('score', 0) * c.get('credit', 0)
        total2 += c.get('credit', 0)
    return total1 / total2 if total2 > 0 else None

class Handler(http.server.SimpleHTTPRequestHandler):
    def _send_json(self, obj, code=200):
        s = json.dumps(obj, ensure_ascii=False).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(s)))
        self.end_headers()
        self.wfile.write(s)

    def do_GET(self):
        parsed = urlparse(self.path)
        # API: students list
        if parsed.path == '/api/students':
            students = load_students()
            for s in students:
                s['gpa'] = calc_gpa(s.get('courses', []))
            self._send_json(students)
            return
        # API: rank
        if parsed.path == '/api/rank':
            students = load_students()
            students_sorted = sorted(students, key=lambda s: (calc_gpa(s.get('courses', [])) or -1), reverse=True)
            out = []
            for i,s in enumerate(students_sorted, start=1):
                out.append({'rank': i, 'id': s['id'], 'name': s.get('name',''), 'gpa': calc_gpa(s.get('courses', []))})
            self._send_json(out)
            return
        # API: export students.csv (Excel-friendly UTF-8 CSV)
        if parsed.path == '/api/export':
            # Export students as CSV (UTF-8 with BOM so Excel recognizes UTF-8)
            students = load_students()
            # CSV header for student basic info
            lines = []
            header = ['id','name','gender','age','college','classnum','plcstatus','phone','province','parphone']
            lines.append(','.join(header))
            for s in students:
                row = [str(s.get(k,'')) for k in header]
                # escape double quotes and wrap fields containing commas
                def esc(field):
                    f = str(field)
                    if ',' in f or '"' in f or '\n' in f:
                        f = '"' + f.replace('"','""') + '"'
                    return f
                row = [esc(x) for x in row]
                lines.append(','.join(row))
            csv_bytes = ('\ufeff' + '\n'.join(lines)).encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'text/csv; charset=utf-8')
            self.send_header('Content-Disposition', 'attachment; filename="students.csv"')
            self.send_header('Content-Length', str(len(csv_bytes)))
            self.end_headers()
            self.wfile.write(csv_bytes)
            return

        if parsed.path == '/api/export_scores':
            # Export scores as CSV: id,course,credit,score
            students = load_students()
            lines = []
            header = ['id','course','credit','score']
            lines.append(','.join(header))
            for s in students:
                sid = s.get('id','')
                for c in s.get('courses',[]):
                    row = [sid, c.get('name',''), str(c.get('credit','')), str(c.get('score',''))]
                    def esc(field):
                        f = str(field)
                        if ',' in f or '"' in f or '\n' in f:
                            f = '"' + f.replace('"','""') + '"'
                        return f
                    row = [esc(x) for x in row]
                    lines.append(','.join(row))
            csv_bytes = ('\ufeff' + '\n'.join(lines)).encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'text/csv; charset=utf-8')
            self.send_header('Content-Disposition', 'attachment; filename="scores.csv"')
            self.send_header('Content-Length', str(len(csv_bytes)))
            self.end_headers()
            self.wfile.write(csv_bytes)
            return

        # API: export as real Excel .xlsx (students)
        if parsed.path == '/api/export_xlsx':
            try:
                from openpyxl import Workbook
            except Exception as e:
                self._send_json({'error': 'openpyxl required: ' + str(e)}, 500)
                return
            students = load_students()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Students'
            header = ['id','name','gender','age','college','classnum','plcstatus','phone','province','parphone']
            # append header and style
            ws.append(header)
            from openpyxl.styles import Font
            bold = Font(bold=True)
            for col_cell in ws[1]:
                col_cell.font = bold
            for s in students:
                row = [s.get(k,'') for k in header]
                ws.append(row)
            # auto column width
            for column_cells in ws.columns:
                length = max((len(str(cell.value)) for cell in column_cells), default=0)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 50)
            # freeze header
            ws.freeze_panes = 'A2'
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            data = bio.read()
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="students.xlsx"')
            self.send_header('Content-Length', str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        # API: export scores as real Excel .xlsx
        if parsed.path == '/api/export_scores_xlsx':
            try:
                from openpyxl import Workbook
            except Exception as e:
                self._send_json({'error': 'openpyxl required: ' + str(e)}, 500)
                return
            students = load_students()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Scores'
            header = ['id','course','credit','score']
            ws.append(header)
            from openpyxl.styles import Font
            bold = Font(bold=True)
            for col_cell in ws[1]:
                col_cell.font = bold
            for s in students:
                sid = s.get('id','')
                for c in s.get('courses',[]):
                    ws.append([sid, c.get('name',''), c.get('credit',''), c.get('score','')])
            # auto column width
            for column_cells in ws.columns:
                length = max((len(str(cell.value)) for cell in column_cells), default=0)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 50)
            ws.freeze_panes = 'A2'
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            data = bio.read()
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="scores.xlsx"')
            self.send_header('Content-Length', str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        # API: save xlsx to server and return JSON with URL
        if parsed.path == '/api/export_xlsx_save':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
            except Exception as e:
                self._send_json({'error': 'openpyxl required: ' + str(e)}, 500)
                return
            students = load_students()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Students'
            header = ['id','name','gender','age','college','classnum','plcstatus','phone','province','parphone']
            ws.append(header)
            bold = Font(bold=True)
            for col_cell in ws[1]:
                col_cell.font = bold
            for s in students:
                row = [s.get(k,'') for k in header]
                ws.append(row)
            for column_cells in ws.columns:
                length = max((len(str(cell.value)) for cell in column_cells), default=0)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 50)
            ws.freeze_panes = 'A2'
            # save to exports dir
            import datetime
            fname = f"students_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            fpath = os.path.join(EXPORT_DIR, fname)
            wb.save(fpath)
            # update metadata
            try:
                with open(EXPORT_METADATA, 'r', encoding='utf-8') as mf:
                    meta = json.load(mf)
            except Exception:
                meta = []
            try:
                mtime = int(os.stat(fpath).st_mtime)
            except Exception:
                mtime = None
            meta.append({'name': fname, 'mtime': mtime, 'saved_time': int(datetime.datetime.now().timestamp())})
            try:
                with open(EXPORT_METADATA, 'w', encoding='utf-8') as mf:
                    json.dump(meta, mf, ensure_ascii=False, indent=2)
            except Exception:
                pass
            url = f'/exports/{fname}'
            self._send_json({'ok': True, 'url': url, 'name': fname})
            return

        if parsed.path == '/api/export_scores_xlsx_save':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
            except Exception as e:
                self._send_json({'error': 'openpyxl required: ' + str(e)}, 500)
                return
            students = load_students()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Scores'
            header = ['id','course','credit','score']
            ws.append(header)
            bold = Font(bold=True)
            for col_cell in ws[1]:
                col_cell.font = bold
            for s in students:
                sid = s.get('id','')
                for c in s.get('courses',[]):
                    ws.append([sid, c.get('name',''), c.get('credit',''), c.get('score','')])
            for column_cells in ws.columns:
                length = max((len(str(cell.value)) for cell in column_cells), default=0)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 50)
            ws.freeze_panes = 'A2'
            import datetime
            fname = f"scores_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            fpath = os.path.join(EXPORT_DIR, fname)
            wb.save(fpath)
            # update metadata
            try:
                with open(EXPORT_METADATA, 'r', encoding='utf-8') as mf:
                    meta = json.load(mf)
            except Exception:
                meta = []
            try:
                mtime = int(os.stat(fpath).st_mtime)
            except Exception:
                mtime = None
            meta.append({'name': fname, 'mtime': mtime, 'saved_time': int(datetime.datetime.now().timestamp())})
            try:
                with open(EXPORT_METADATA, 'w', encoding='utf-8') as mf:
                    json.dump(meta, mf, ensure_ascii=False, indent=2)
            except Exception:
                pass
            url = f'/exports/{fname}'
            self._send_json({'ok': True, 'url': url, 'name': fname})
            return

        # API: list saved exports
        if parsed.path == '/api/exports':
            # return export history from metadata (does not delete files)
            try:
                with open(EXPORT_METADATA, 'r', encoding='utf-8') as f:
                    meta = json.load(f)
            except Exception:
                meta = []
            files = []
            for entry in sorted(meta, key=lambda e: e.get('saved_time', 0), reverse=True):
                name = entry.get('name')
                url = f'/exports/{name}'
                mtime = None
                fpath = os.path.join(EXPORT_DIR, name)
                if os.path.exists(fpath):
                    try:
                        mtime = int(os.stat(fpath).st_mtime)
                    except Exception:
                        mtime = entry.get('mtime')
                else:
                    # file missing on disk; keep metadata but mark mtime as None
                    mtime = entry.get('mtime')
                files.append({'name': name, 'url': url, 'mtime': mtime})
            self._send_json({'files': files})
            return

        # Serve static files / index
        if self.path == '/' or self.path == '/index.html':
            self.path = '/templates/index.html'
        return http.server.SimpleHTTPRequestHandler.do_GET(self)

    def do_POST(self):
        parsed = urlparse(self.path)
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8') if length else ''
        data = json.loads(body) if body else {}
        if parsed.path == '/api/students':
            students = load_students()
            if any(s['id'] == data['id'] for s in students):
                self._send_json({'error':'学号已存在'}, 400); return
            data['courses'] = []
            students.append(data); save_students(students)
            self._send_json({'ok':True}); return
        if parsed.path.startswith('/api/students/') and parsed.path.endswith('/scores'):
            from urllib.parse import unquote
            # extract and URL-decode student id
            prefix = '/api/students/'
            sid = parsed.path[len(prefix):].rsplit('/',1)[0]
            sid = unquote(sid)
            students = load_students()
            for s in students:
                if s['id'] == sid:
                    for c in s.setdefault('courses', []):
                        if c['name'] == data['name']:
                            c['credit'] = data.get('credit', c.get('credit',0))
                            c['score'] = data.get('score', c.get('score',0))
                            save_students(students); self._send_json({'ok':True}); return
                    s.setdefault('courses', []).append({'name':data['name'], 'credit':data.get('credit',0), 'score':data.get('score',0)})
                    save_students(students); self._send_json({'ok':True}); return
            self._send_json({'error':'未找到学号'},404); return

    def do_PUT(self):
        parsed = urlparse(self.path)
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8') if length else ''
        data = json.loads(body) if body else {}
        if parsed.path.startswith('/api/students/'):
            from urllib.parse import unquote
            prefix = '/api/students/'
            sid = parsed.path[len(prefix):].split('/')[0]
            sid = unquote(sid)
            students = load_students()
            for s in students:
                if s['id'] == sid:
                    for k in ['name','gender','age','college','classnum','plcstatus','phone','province','parphone']:
                        if k in data: s[k] = data[k]
                    save_students(students); self._send_json({'ok':True}); return
            self._send_json({'error':'未找到学号'},404); return

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith('/api/students/'):
            from urllib.parse import unquote
            prefix = '/api/students/'
            sid = parsed.path[len(prefix):].split('/')[0]
            sid = unquote(sid)
            students = load_students(); new=[s for s in students if s['id']!=sid]
            if len(new)==len(students): self._send_json({'error':'未找到学号'},404); return
            save_students(new); self._send_json({'ok':True}); return
        if parsed.path.startswith('/api/exports/'):
            # delete saved export file
            # extract filename after the prefix '/api/exports/' and URL-decode it
            from urllib.parse import unquote
            prefix = '/api/exports/'
            if not parsed.path.startswith(prefix):
                self._send_json({'error': 'bad request'}, 400); return
            fname = unquote(parsed.path[len(prefix):])
            if not fname:
                self._send_json({'error': 'bad request'}, 400); return
            # prevent directory traversal
            if '..' in fname or '/' in fname:
                self._send_json({'error':'invalid filename'}, 400); return
            # New behavior: only remove the history metadata entry, keep the actual file on disk.
            try:
                with open(EXPORT_METADATA, 'r', encoding='utf-8') as mf:
                    meta = json.load(mf)
            except Exception:
                meta = []
            new_meta = [e for e in meta if e.get('name') != fname]
            if len(new_meta) == len(meta):
                # not found in metadata
                self._send_json({'error':'not found in history'}, 404); return
            try:
                with open(EXPORT_METADATA, 'w', encoding='utf-8') as mf:
                    json.dump(new_meta, mf, ensure_ascii=False, indent=2)
            except Exception as e:
                self._send_json({'error': str(e)}, 500); return
            self._send_json({'ok': True})
            return


if __name__ == '__main__':
    START_PORT = 5000
    MAX_TRIES = 11  # try ports 5000 .. 5010
    os.chdir(ROOT)
    ensure_data_dir()
    socketserver.ThreadingTCPServer.allow_reuse_address = True
    httpd = None
    for i in range(MAX_TRIES):
        port = START_PORT + i
        try:
            httpd = socketserver.ThreadingTCPServer(('0.0.0.0', port), Handler)
            print(f"Serving at http://127.0.0.1:{port}/")
            break
        except OSError as e:
            print(f"Port {port} unavailable: {e}. Trying next port...")
            httpd = None
            continue
    if httpd is None:
        print(f"Failed to bind any port in range {START_PORT}-{START_PORT+MAX_TRIES-1}. Exiting.")
    else:
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print('shutting down')
        finally:
            try:
                httpd.server_close()
            except Exception:
                pass
